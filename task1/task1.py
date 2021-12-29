import openpyxl
import json


book = openpyxl.Workbook()
del book['Sheet']


def parse_json_to_sheet(file_name, sheet_num):

    sheet1 = book.create_sheet(file_name[:-5], sheet_num)

    with open(file_name, encoding='utf-8') as file:
        data = json.load(file)

    sheet1['A1'] = 1
    rows = {}
    d = {}

    # data extraction (columns)
    for key, value in data.items():
        for v in value:
            x = 'Y' + v['properties']['Y'] + 'X' + v['properties']['X']
            if key == 'headers':
                d[x] = v['properties']['QuickInfo']
            elif key == 'values':
                d[x] = v['properties']['Text'].strip()

    # rows generation
    for k, v in d.items():
        index = k[1]
        if index not in rows:
            rows[index] = []
        rows[index].append({k[2:]: v})

    # write data
    i = 1
    for row, column in rows.items():
        for col in column:
            for k, v in col.items():
                if k == 'X28':
                    sheet1.cell(row=i, column=1).value = v
                elif k == 'X39':
                    sheet1.cell(row=i, column=2).value = v
                elif k == 'X62':
                    sheet1.cell(row=i, column=3).value = v
                elif k == 'X120':
                    sheet1.cell(row=i, column=4).value = v
        i += 1


parse_json_to_sheet('test1.json', 0)
parse_json_to_sheet('test2.json', 1)
parse_json_to_sheet('test3.json', 2)

book.save('task1.xlsx')
book.close()
